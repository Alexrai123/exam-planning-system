"""
Script to create a new exam template with faculty and specialization information
"""
import sys
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Add the parent directory to the Python path
sys.path.append(str(Path(__file__).parent.parent))

def create_new_exam_template():
    """
    Create a new exam template with faculty and specialization information
    """
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exam Data"
        
        # Define headers with faculty and specialization information
        headers = [
            "Faculty Name",
            "Faculty Short Name",
            "Specialization Name",
            "Specialization Short Name",
            "Course ID",
            "Course Name", 
            "Year",
            "Semester",
            "Professor Name", 
            "Date (YYYY-MM-DD)", 
            "Time (HH:MM)", 
            "Room", 
            "Group",
            "Status",
            "Professor Agreement"
        ]
        
        # Define column widths
        column_widths = [30, 15, 30, 15, 10, 30, 10, 10, 30, 20, 15, 20, 15, 15, 20]
        
        # Set column widths
        for i, width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i+1)].width = width
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add example data
        example_data = [
            "Facultatea de Inginerie Electrica si Stiinta Calculatoarelor",
            "FIESC",
            "Calculatoare",
            "CALC",
            "369",
            "Algebra liniara, geometrie analitica si diferentiala",
            "1",
            "1",
            "N/A",
            "2025-06-15",
            "10:00",
            "C11",
            "CALC1A",
            "PROPOSED",
            "FALSE"
        ]
        
        # Add example row
        for col_num, value in enumerate(example_data, 1):
            cell = ws.cell(row=2, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Add instructions
        ws.cell(row=4, column=1, value="Instructions:")
        ws.cell(row=5, column=1, value="1. Fill in the exam data following the example above")
        ws.cell(row=6, column=1, value="2. Do not modify the header row")
        ws.cell(row=7, column=1, value="3. Status must be one of: PROPOSED, CONFIRMED, CANCELED")
        ws.cell(row=8, column=1, value="4. Professor Agreement must be TRUE or FALSE")
        ws.cell(row=9, column=1, value="5. Faculty, Specialization, and Course ID must match existing records")
        ws.cell(row=10, column=1, value="6. Save the file as Excel (.xlsx) before uploading")
        
        # Create output directories
        public_dir = Path(__file__).parent.parent / "app" / "public"
        public_dir.mkdir(exist_ok=True)
        
        templates_dir = public_dir / "templates"
        templates_dir.mkdir(exist_ok=True)
        
        # Save the template to multiple locations to ensure it's available
        template_paths = [
            Path(__file__).parent.parent / "static" / "templates" / "exam_template.xlsx",
            templates_dir / "exam_template.xlsx",
            Path(__file__).parent.parent / "app" / "templates" / "exam_template.xlsx"
        ]
        
        for template_path in template_paths:
            # Create parent directories if they don't exist
            template_path.parent.mkdir(exist_ok=True, parents=True)
            
            # Save the template
            wb.save(template_path)
            print(f"Template saved to: {template_path}")
        
        print("\nNew exam template created successfully with faculty and specialization information.")
        print("Please restart the backend server for changes to take effect.")
        
    except Exception as e:
        print(f"Error creating new exam template: {e}")

if __name__ == "__main__":
    create_new_exam_template()
