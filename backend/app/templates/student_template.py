"""
Generate Excel template for student data
"""
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_student_template():
    """
    Generate an Excel template for student data
    """
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Data"
    
    # Define headers
    headers = [
        "Student Name", 
        "Student Email", 
        "Group", 
        "Year", 
        "Specialization", 
        "Faculty",
        "Is Group Leader (TRUE/FALSE)"
    ]
    
    # Define column widths
    column_widths = [30, 30, 15, 10, 30, 30, 25]
    
    # Set column widths
    for i, width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = width
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Add headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add example data
    example_data = [
        "John Smith",
        "john.smith@student.usv.ro",
        "Group 1",
        "1",
        "Computer Science",
        "Faculty of Computer Science",
        "TRUE"
    ]
    
    # Add example row
    for col_num, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_num, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Add instructions
    ws.cell(row=4, column=1, value="Instructions:")
    ws.cell(row=5, column=1, value="1. Fill in the student data following the example above")
    ws.cell(row=6, column=1, value="2. Do not modify the header row")
    ws.cell(row=7, column=1, value="3. Student email should be in the format: firstname.lastname@student.usv.ro")
    ws.cell(row=8, column=1, value="4. Is Group Leader must be TRUE or FALSE")
    ws.cell(row=9, column=1, value="5. Save the file as Excel (.xlsx) before uploading")
    
    # Save to a BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
