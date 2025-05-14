"""
Generate Excel template for exam data
"""
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_exam_template():
    """
    Generate an Excel template for exam data with the new structure
    """
    from . import REORDERED_TEMPLATE_PATH
    
    # Check if the reordered template exists
    if REORDERED_TEMPLATE_PATH.exists():
        print(f"Using existing reordered template: {REORDERED_TEMPLATE_PATH}")
        # Read the existing template file
        with open(REORDERED_TEMPLATE_PATH, "rb") as f:
            content = f.read()
        output = io.BytesIO(content)
        output.seek(0)
        return output
        
    print("Reordered template not found, generating dynamically")
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Exam Data"
    
    # Define headers in the requested order
    headers = [
        "Faculty",
        "Specialization",
        "Year",
        "Group",
        "Course", 
        "Professor", 
        "Date (YYYY-MM-DD)", 
        "Time (HH:MM)", 
        "Room", 
        "Status",
        "Professor Agreement"
    ]
    
    # Define column widths
    column_widths = [30, 30, 10, 15, 30, 30, 20, 15, 15, 15, 20]
    
    # Set column widths
    for i, width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = width
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Add headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add example data
    example_data = [
        "Facultatea de Inginerie Electrica si Stiinta Calculatoarelor",
        "Calculatoare",
        "1",
        "CALC1A",
        "Algebra liniara, geometrie analitica si diferentiala",
        "Prof. Dr. Maria Popescu",
        "2025-06-15",
        "10:00",
        "C11",
        "PROPOSED",
        "FALSE"
    ]
    
    # Add example row
    for col_num, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_num, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Save to a BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
